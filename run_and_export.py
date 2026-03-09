#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
run_and_export.py — Run pricer + export structured Excel for Power BI
=====================================================================
Usage:
  python run_and_export.py
  python run_and_export.py --config config/config.yaml
  python run_and_export.py --config config/config.yaml --output output/pbi_data.xlsx

The output Excel has clean table-formatted sheets that Power BI can read directly.
In Power BI: Get Data → Excel → select the file → load all tables.
"""

import os, sys, argparse
from datetime import datetime

src_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "src")
if src_dir not in sys.path:
    sys.path.insert(0, src_dir)

import yaml
import numpy as np
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from pricer import BermudanPricer
from bbg_fetcher import fetch_all


def run_pricer(config_path):
    """Run pricer from config file, return pricer object."""
    config_dir = os.path.dirname(os.path.abspath(config_path))
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    mkt = fetch_all(cfg, config_dir=config_dir)
    pricer = BermudanPricer(cfg, mkt)
    pricer.setup()
    pricer.calibrate()
    pricer.compute_greeks()
    return pricer, cfg


def export_pbi_excel(pricer, cfg, output_path):
    """Export results in a Power BI-friendly Excel format."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font_w = Font(bold=True, color="FFFFFF", size=11)

    def make_table(ws, name, ref):
        """Add an Excel Table (structured reference) to a sheet."""
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tab)

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 1: Summary — single row with all key metrics
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"

    g = pricer.greeks
    bps_leg = abs(float(pricer.swap.fixedLegBPS()))
    yv = pricer.npv / bps_leg if bps_leg else 0

    headers = [
        "ValDate", "Notional", "Strike_pct", "Direction", "SwapStart", "SwapEnd",
        "ATM_pct", "Moneyness_bp", "NPV", "YieldValue_bp", "Premium_pct",
        "Und_Premium_pct", "Und_NPV",
        "sigma_ATM_bp", "delta_spread_bp", "sigma_total_bp", "a",
        "DV01", "Gamma_1bp", "Vega_1bp", "Theta_1d", "Delta", "Und_DV01",
    ]
    values = [
        str(pricer.val_date),
        pricer.notional,
        pricer.strike * 100,
        "Receiver" if pricer.is_receiver else "Payer",
        str(pricer.swap_start),
        str(pricer.swap_end),
        pricer.fair_rate * 100,
        (pricer.strike - pricer.fair_rate) * 10000,
        pricer.npv,
        yv,
        pricer.npv / pricer.notional * 100,
        pricer.underlying_npv / pricer.notional * 100,
        pricer.underlying_npv,
        pricer.sigma_atm * 10000,
        pricer.delta_spread * 10000,
        pricer.sigma_total * 10000,
        pricer.a,
        g["dv01"], g["gamma_1bp"], g["vega_1bp"], g["theta_1d"],
        g["delta_hedge"], g["underlying_dv01"],
    ]

    for j, h in enumerate(headers, 1):
        ws.cell(1, j, h)
    for j, v in enumerate(values, 1):
        ws.cell(2, j, v)

    make_table(ws, "tblSummary", f"A1:{chr(64+len(headers))}2")

    # Auto-width
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or "")) + 3, 12
        )

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 2: BBG Comparison — one row per metric
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("BBG_Comparison")
    ws2.append(["Metric", "Bloomberg", "QuantLib", "Diff", "Diff_pct"])

    bench = cfg.get("benchmark", {})
    bbg_npv = float(bench.get("npv", 0) or 0)

    comparisons = [
        ("NPV", bbg_npv, pricer.npv),
        ("ATM_Strike_pct", bench.get("atm_strike"), pricer.fair_rate * 100),
        ("YieldValue_bp", bench.get("yield_value_bp"), yv),
        ("Und_Premium_pct", bench.get("underlying_premium"), pricer.underlying_npv / pricer.notional * 100),
        ("Premium_pct", bench.get("premium"), pricer.npv / pricer.notional * 100),
        ("DV01", bench.get("dv01"), g["dv01"]),
        ("Gamma_1bp", bench.get("gamma_1bp"), g["gamma_1bp"]),
        ("Vega_1bp", bench.get("vega_1bp"), g["vega_1bp"]),
        ("Theta_1d", bench.get("theta_1d"), g["theta_1d"]),
    ]

    for name, bbg, ql_val in comparisons:
        if bbg is not None:
            bbg = float(bbg)
            diff = ql_val - bbg
            pct = (diff / abs(bbg) * 100) if bbg != 0 else 0
            ws2.append([name, bbg, ql_val, diff, pct])
        else:
            ws2.append([name, None, ql_val, None, None])

    make_table(ws2, "tblComparison", f"A1:E{len(comparisons)+1}")
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 18

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 3: Greeks — structured for easy charting
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Greeks")
    ws3.append(["Greek", "Value", "BBG", "Diff", "Diff_pct"])

    greek_rows = [
        ("DV01", g["dv01"], bench.get("dv01")),
        ("Gamma_1bp", g["gamma_1bp"], bench.get("gamma_1bp")),
        ("Vega_1bp", g["vega_1bp"], bench.get("vega_1bp")),
        ("Theta_1d", g["theta_1d"], bench.get("theta_1d")),
        ("Delta", g["delta_hedge"], None),
        ("Und_DV01", g["underlying_dv01"], None),
    ]
    for name, val, bbg in greek_rows:
        if bbg is not None:
            bbg = float(bbg)
            diff = val - bbg
            pct = (diff / abs(bbg) * 100) if bbg != 0 else 0
            ws3.append([name, val, bbg, diff, pct])
        else:
            ws3.append([name, val, None, None, None])

    make_table(ws3, "tblGreeks", f"A1:E{len(greek_rows)+1}")
    for col in ws3.columns:
        ws3.column_dimensions[col[0].column_letter].width = 16

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 4: Curve
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Curve")
    ws4.append(["Date", "DiscountFactor"])
    for d, df in pricer.mkt["curve"]:
        ws4.append([str(d), float(df)])
    make_table(ws4, "tblCurve", f"A1:B{len(pricer.mkt['curve'])+1}")
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 16

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 5: Vol Surface
    # ══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("VolSurface")
    vsd = cfg.get("vol_surface_data", {})
    tnr_labels = vsd.get("tenor_labels", [])
    exp_labels = vsd.get("expiry_labels", [])
    raw_vol = pricer.vol_mat * 1000.0  # back to BPx10

    ws5.append(["Expiry"] + tnr_labels)
    for i, exp in enumerate(exp_labels):
        row_data = [exp] + [float(raw_vol[i, j]) for j in range(raw_vol.shape[1])]
        ws5.append(row_data)

    last_col = chr(65 + len(tnr_labels)) if len(tnr_labels) < 26 else "P"
    make_table(ws5, "tblVol", f"A1:{last_col}{len(exp_labels)+1}")

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 6: Run Log (metadata)
    # ══════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("RunLog")
    ws6.append(["Parameter", "Value"])
    log_data = [
        ("RunTimestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("ConfigFile", "config.yaml"),
        ("Model", "HW1F"),
        ("a_fixed_or_calibrated", "calibrated" if pricer.calib_a else "fixed"),
        ("a_value", pricer.a),
        ("sigma_ATM_bp", pricer.sigma_atm * 10000),
        ("delta_spread_bp", pricer.delta_spread * 10000),
        ("sigma_total_bp", pricer.sigma_total * 10000),
        ("FDM_grid", f"{pricer.fdm_t}x{pricer.fdm_x}"),
        ("NumExerciseDates", len(pricer.ex_dates)),
        ("BBG_NPV_target", pricer.bbg_npv),
        ("NPV_match_pct", f"{(pricer.npv - pricer.bbg_npv) / pricer.bbg_npv * 100:.6f}%" if pricer.bbg_npv else "N/A"),
    ]
    for k, v in log_data:
        ws6.append([k, v])
    make_table(ws6, "tblRunLog", f"A1:B{len(log_data)+1}")
    ws6.column_dimensions["A"].width = 25
    ws6.column_dimensions["B"].width = 30

    # Save
    wb.save(output_path)
    print(f"\n  ✓ Power BI Excel exported to: {output_path}")
    print(f"    Tables: tblSummary, tblComparison, tblGreeks, tblCurve, tblVol, tblRunLog")
    print(f"\n  In Power BI Desktop:")
    print(f"    1. Get Data → Excel Workbook → select '{os.path.basename(output_path)}'")
    print(f"    2. Check all 6 tables → Load")
    print(f"    3. Build your dashboard!")


def main():
    parser = argparse.ArgumentParser(description="Run pricer + export for Power BI")
    parser.add_argument("--config", default=None, help="Config YAML path")
    parser.add_argument("--output", default=None, help="Output Excel path")
    args = parser.parse_args()

    config_path = args.config
    if config_path is None:
        for c in ["config/config.yaml", "config.yaml"]:
            if os.path.exists(c):
                config_path = c
                break
    if not config_path or not os.path.exists(config_path):
        print("Config not found. Use: python run_and_export.py --config config/config.yaml")
        sys.exit(1)

    output_path = args.output or os.path.join("output", "pbi_data.xlsx")
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    print("=" * 70)
    print("  BERMUDAN SWAPTION PRICER → Power BI Export")
    print("=" * 70)

    pricer, cfg = run_pricer(config_path)
    pricer.print_results()
    export_pbi_excel(pricer, cfg, output_path)
    print("\n✓ Done")


if __name__ == "__main__":
    main()

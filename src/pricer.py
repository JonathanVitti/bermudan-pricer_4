#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
pricer.py — Bermudan Swaption Pricer (v12 hybrid)
===================================================
Usage:
  python pricer.py                        # uses config.yaml in current dir
  python pricer.py --config my_deal.yaml  # custom config file
  python pricer.py --config deal.yaml --output results.xlsx

Reads all parameters from YAML config, fetches market data (Bloomberg or manual),
calibrates HW1F with hybrid approach (σ_ATM + Δσ_spread), computes Greeks.
"""

import argparse
import os
import sys
import math
import numpy as np
from datetime import datetime

try:
    import yaml
except ImportError:
    print("PyYAML required: pip install pyyaml")
    sys.exit(1)

import QuantLib as ql
from scipy.optimize import minimize, brentq

from bbg_fetcher import fetch_all

# ═══════════════════════════════════════════════════════════════════════════
#  CONFIG PARSER
# ═══════════════════════════════════════════════════════════════════════════

def parse_date(s):
    """Parse YYYY-MM-DD string to ql.Date."""
    if isinstance(s, ql.Date):
        return s
    d = datetime.strptime(str(s).strip(), "%Y-%m-%d")
    return ql.Date(d.day, d.month, d.year)


def parse_frequency(s):
    m = str(s).lower().replace(" ", "")
    return {
        "semiannual": ql.Period(6, ql.Months),
        "quarterly":  ql.Period(3, ql.Months),
        "annual":     ql.Period(12, ql.Months),
        "monthly":    ql.Period(1, ql.Months),
    }.get(m, ql.Period(6, ql.Months))


def parse_daycount(s):
    m = str(s).upper().replace(" ", "").replace("/", "")
    return {
        "ACT365": ql.Actual365Fixed(),
        "ACT360": ql.Actual360(),
        "30360":  ql.Thirty360(ql.Thirty360.BondBasis),
    }.get(m, ql.Actual365Fixed())


def parse_direction(s):
    m = str(s).lower().strip()
    if m in ("receiver", "rec", "r"):
        return ql.OvernightIndexedSwap.Receiver
    return ql.OvernightIndexedSwap.Payer


# ═══════════════════════════════════════════════════════════════════════════
#  QUANTLIB HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def get_calendar(ccy="CAD"):
    try:
        return ql.Canada()
    except TypeError:
        return ql.Canada(ql.Canada.Settlement)


def get_index(handle, ccy="CAD"):
    cal = get_calendar(ccy)
    dc = ql.Actual365Fixed()
    if ccy == "CAD":
        if hasattr(ql, "Corra"):
            try: return ql.Corra(handle)
            except: pass
        return ql.OvernightIndex("CORRA", 0, ql.CADCurrency(), cal, dc, handle)
    elif ccy == "USD":
        if hasattr(ql, "Sofr"):
            try: return ql.Sofr(handle)
            except: pass
        return ql.OvernightIndex("SOFR", 0, ql.USDCurrency(), cal, dc, handle)
    else:
        return ql.OvernightIndex("OIS", 0, ql.USDCurrency(), cal, dc, handle)


def build_curve(ref_date, dates, dfs, cal, dc):
    c = ql.DiscountCurve([ref_date] + dates, [1.0] + dfs, dc, cal)
    c.enableExtrapolation()
    return ql.YieldTermStructureHandle(c), c


def make_schedule(start, end, tenor, cal, bdc):
    return ql.Schedule(start, end, tenor, cal, bdc, bdc,
                       ql.DateGeneration.Backward, False)


def make_ois(direction, nom, schedule, rate, index, fixed_dc,
             payment_lag=2, bdc=ql.ModifiedFollowing, cal=None,
             telescopic=False, averaging=ql.RateAveraging.Compound):
    if cal is None:
        cal = get_calendar()
    return ql.OvernightIndexedSwap(
        direction, nom, schedule, rate, fixed_dc, index, 0.0,
        payment_lag, bdc, cal, telescopic, averaging)


def make_swaption(swap, exercise):
    try:
        st = ql.Settlement.Type.Cash
        for m in ("CollateralizedCashPrice", "ParYieldCurve"):
            if hasattr(ql.Settlement, "Method") and hasattr(ql.Settlement.Method, m):
                try: return ql.Swaption(swap, exercise, st, getattr(ql.Settlement.Method, m))
                except: pass
        try: return ql.Swaption(swap, exercise, st)
        except: pass
    except AttributeError: pass
    try: return ql.Swaption(swap, exercise, ql.Settlement.Type.Physical)
    except: return ql.Swaption(swap, exercise)


# ═══════════════════════════════════════════════════════════════════════════
#  PRICING ENGINE
# ═══════════════════════════════════════════════════════════════════════════

def bachelier_receiver(F, K, sigma, T, annuity):
    if sigma <= 0.0 or T <= 0.0:
        return annuity * max(K - F, 0.0)
    std = sigma * math.sqrt(T)
    d = (F - K) / std
    phi = math.exp(-0.5 * d * d) / math.sqrt(2.0 * math.pi)
    Phi_m = 0.5 * (1.0 + math.erf(-d / math.sqrt(2.0)))
    return annuity * ((K - F) * Phi_m + std * phi)


def bachelier_payer(F, K, sigma, T, annuity):
    if sigma <= 0.0 or T <= 0.0:
        return annuity * max(F - K, 0.0)
    std = sigma * math.sqrt(T)
    d = (F - K) / std
    phi = math.exp(-0.5 * d * d) / math.sqrt(2.0 * math.pi)
    Phi = 0.5 * (1.0 + math.erf(d / math.sqrt(2.0)))
    return annuity * ((F - K) * Phi + std * phi)


def vol_interp(T, tenor, vol_mat, expiry_grid, tenor_grid):
    xc = float(np.clip(T, expiry_grid[0], expiry_grid[-1]))
    yc = float(np.clip(tenor, tenor_grid[0], tenor_grid[-1]))
    i1 = min(np.searchsorted(expiry_grid, xc), len(expiry_grid) - 1)
    j1 = min(np.searchsorted(tenor_grid, yc), len(tenor_grid) - 1)
    i0, j0 = max(i1 - 1, 0), max(j1 - 1, 0)
    wx = 0 if expiry_grid[i1] == expiry_grid[i0] else (xc - expiry_grid[i0]) / (expiry_grid[i1] - expiry_grid[i0])
    wy = 0 if tenor_grid[j1] == tenor_grid[j0] else (yc - tenor_grid[j0]) / (tenor_grid[j1] - tenor_grid[j0])
    return float((1-wy)*((1-wx)*vol_mat[i0,j0]+wx*vol_mat[i1,j0]) +
                 wy*((1-wx)*vol_mat[i0,j1]+wx*vol_mat[i1,j1]))


class BermudanPricer:
    """Full Bermudan swaption pricer with hybrid calibration."""

    def __init__(self, cfg, market_data):
        self.cfg = cfg
        self.mkt = market_data

        # Parse deal
        deal = cfg["deal"]
        self.val_date    = parse_date(deal["valuation_date"])
        self.notional    = float(deal["notional"])
        self.strike      = float(deal["strike"]) / 100.0
        self.direction   = parse_direction(deal.get("direction", "Receiver"))
        self.swap_start  = parse_date(deal["swap_start"])
        self.swap_end    = parse_date(deal["swap_end"])
        self.fixed_tenor = parse_frequency(deal.get("fixed_frequency", "SemiAnnual"))
        self.fixed_dc    = parse_daycount(deal.get("day_count", "ACT/365"))
        self.payment_lag = int(deal.get("payment_lag", 2))
        self.ccy         = deal.get("currency", "CAD")
        self.is_receiver = (self.direction == ql.OvernightIndexedSwap.Receiver)

        # Model
        model = cfg.get("model", {})
        self.a         = float(model.get("mean_reversion", 0.03))
        self.calib_a   = bool(model.get("calibrate_a", False))
        self.fdm_t     = int(model.get("fdm_time_grid", 300))
        self.fdm_x     = int(model.get("fdm_space_grid", 300))

        # Greeks config
        gk = cfg.get("greeks", {})
        self.dv01_bp    = float(gk.get("dv01_bump_bp", 1.0))
        self.gamma_bp   = float(gk.get("gamma_bump_bp", 1.0))
        self.vega_bp    = float(gk.get("vega_bump_bp", 1.0))
        self.do_theta   = bool(gk.get("compute_theta", True))
        self.theta_ann  = gk.get("theta_annualization", "365/252")

        # Benchmark
        bench = cfg.get("benchmark", {})
        self.bbg_npv = float(market_data["bbg_npv"])
        self.bbg = bench

        # QL setup
        self.cal = get_calendar(self.ccy)
        self.bdc = ql.ModifiedFollowing
        self.dc  = ql.Actual365Fixed()

        # Vol surface
        self.vol_mat    = market_data["vol_surface"] / 1000.0  # BPx10 → decimal
        self.exp_grid   = market_data["expiry_grid"]
        self.tnr_grid   = market_data["tenor_grid"]

    def setup(self):
        """Initialize QuantLib objects."""
        ql.Settings.instance().evaluationDate = self.val_date

        # Curve — parse, filter dates <= val_date, sort, dedup
        raw_pairs = []
        for d, df in self.mkt["curve"]:
            try:
                dt = parse_date(d)
                raw_pairs.append((dt, float(df)))
            except Exception:
                continue

        # Filter out dates on or before valuation date
        pairs = [(dt, df) for dt, df in raw_pairs if dt.serialNumber() > self.val_date.serialNumber()]

        if not pairs:
            raise ValueError(f"No curve nodes after valuation date {self.val_date}. "
                             f"Check that curve dates are after {self.val_date}.")

        # Sort by date
        pairs.sort(key=lambda x: x[0].serialNumber())

        # Dedup (keep last value for same date)
        dedup = []
        for dt, df in pairs:
            if dedup and dt.serialNumber() == dedup[-1][0].serialNumber():
                dedup[-1] = (dt, df)
            else:
                dedup.append((dt, df))

        self.node_dates = [d for d, _ in dedup]
        self.node_dfs   = [df for _, df in dedup]

        # Validate DFs
        for i, (dt, df) in enumerate(zip(self.node_dates, self.node_dfs)):
            if df <= 0:
                raise ValueError(f"Discount factor <= 0 at {dt}: {df}")
            if df > 1.05:  # small tolerance for very short dates
                print(f"  [WARNING] DF > 1.05 at {dt}: {df} — check curve data")
        for i in range(1, len(self.node_dfs)):
            if self.node_dfs[i] > self.node_dfs[i-1] + 1e-6:
                print(f"  [WARNING] Non-monotone DFs: {self.node_dates[i-1]}={self.node_dfs[i-1]:.6f} → {self.node_dates[i]}={self.node_dfs[i]:.6f}")

        self.yts_h, self.yts_c = build_curve(self.val_date, self.node_dates, self.node_dfs,
                                              self.cal, self.dc)

        # Index
        self.index = get_index(self.yts_h, self.ccy)

        # Schedule
        self.schedule = make_schedule(self.swap_start, self.swap_end,
                                       self.fixed_tenor, self.cal, self.bdc)

        # Exercise dates
        ex_cfg = self.cfg.get("exercise", {})
        if ex_cfg.get("mode", "auto") == "custom" and "custom_dates" in ex_cfg:
            self.ex_dates = [parse_date(d) for d in ex_cfg["custom_dates"]]
        else:
            # Auto: all schedule dates except last (= swap end)
            all_dates = list(self.schedule)
            self.ex_dates = [d for d in all_dates[:-1] if d >= self.swap_start]

        # Underlying swap
        self.swap = make_ois(self.direction, self.notional, self.schedule,
                             self.strike, self.index, self.fixed_dc,
                             self.payment_lag, self.bdc, self.cal)
        self.swap.setPricingEngine(ql.DiscountingSwapEngine(self.yts_h))
        self.fair_rate = float(self.swap.fairRate())
        self.underlying_npv = float(self.swap.NPV())

    def _build_berm(self, h=None, index=None, schedule=None):
        h = h or self.yts_h
        index = index or self.index
        schedule = schedule or self.schedule
        s = make_ois(self.direction, self.notional, schedule, self.strike,
                     index, self.fixed_dc, self.payment_lag, self.bdc, self.cal)
        s.setPricingEngine(ql.DiscountingSwapEngine(h))
        return s, make_swaption(s, ql.BermudanExercise(self.ex_dates))

    def _price_berm(self, h, swpt, sigma):
        hw = ql.HullWhite(h, self.a, sigma)
        swpt.setPricingEngine(ql.FdHullWhiteSwaptionEngine(hw, self.fdm_t, self.fdm_x))
        return float(swpt.NPV())

    def _bachelier(self, fwd, vol, T, ann):
        if self.is_receiver:
            return bachelier_receiver(fwd, self.strike, vol, T, ann)
        else:
            return bachelier_payer(fwd, self.strike, vol, T, ann)

    def _build_basket(self, h=None, index=None, vol_bump_bp=0.0):
        h = h or self.yts_h
        index = index or self.index
        sd = list(self.schedule)
        basket = []
        for ex in self.ex_dates:
            if ex in sd:
                sub = sd[sd.index(ex):]
                try:
                    ss = ql.Schedule(sub, self.cal, self.bdc)
                    if len(list(ss)) != len(sub):
                        ss = make_schedule(ex, self.swap_end, self.fixed_tenor, self.cal, self.bdc)
                except:
                    ss = make_schedule(ex, self.swap_end, self.fixed_tenor, self.cal, self.bdc)
            else:
                ss = make_schedule(ex, self.swap_end, self.fixed_tenor, self.cal, self.bdc)

            s0 = make_ois(self.direction, self.notional, ss, 0.0, index,
                          self.fixed_dc, self.payment_lag, self.bdc, self.cal)
            s0.setPricingEngine(ql.DiscountingSwapEngine(h))
            fwd = float(s0.fairRate())

            sk = make_ois(self.direction, self.notional, ss, self.strike, index,
                          self.fixed_dc, self.payment_lag, self.bdc, self.cal)
            sk.setPricingEngine(ql.DiscountingSwapEngine(h))
            ann = abs(float(sk.fixedLegBPS())) / 1e-4

            T    = self.dc.yearFraction(self.val_date, ex)
            tenY = self.dc.yearFraction(ex, self.swap_end)
            vol  = vol_interp(T, tenY, self.vol_mat, self.exp_grid, self.tnr_grid)
            vol += vol_bump_bp / 10000.0

            swpt = ql.Swaption(sk, ql.EuropeanExercise(ex))
            mkt  = self._bachelier(fwd, vol, T, ann)
            basket.append(dict(fwd=fwd, vol=vol, T=T, ann=ann, swpt=swpt, mkt=mkt))
        return basket

    def _calib_sigma_atm(self, h, basket):
        """Calibrate σ only (a is fixed at self.a).
        Uses JamshidianSwaptionEngine (analytic) for European swaptions —
        no grid dependency, stable calibration.
        """
        def obj(x):
            sigma = 1e-8 + math.exp(float(x[0]))
            hw = ql.HullWhite(h, self.a, sigma)
            eng = ql.JamshidianSwaptionEngine(hw)
            return sum((float(it["swpt"].setPricingEngine(eng) or it["swpt"].NPV()) - it["mkt"])**2
                       / max(1, abs(it["mkt"])) for it in basket)
        res = minimize(obj, [math.log(0.005)], method="Nelder-Mead",
                       options={"maxiter": 500, "xatol": 1e-8, "fatol": 1e-8})
        if not res.success:
            print(f"  [WARNING] σ_ATM calibration did not converge: {res.message}")
        return 1e-8 + math.exp(float(res.x[0]))

    def _calib_joint(self, h, basket):
        """Calibrate (a, σ) jointly on European basket.
        a is bounded to [0.001, 0.50] to avoid degenerate solutions.
        """
        A_MIN, A_MAX = 0.001, 0.50

        def _logit(a):
            """Map a in [A_MIN, A_MAX] → R."""
            t = (a - A_MIN) / (A_MAX - A_MIN)
            t = max(1e-10, min(1 - 1e-10, t))
            return math.log(t / (1 - t))

        def _inv_logit(x):
            """Map R → a in [A_MIN, A_MAX]."""
            t = 1.0 / (1.0 + math.exp(-float(x)))
            return A_MIN + t * (A_MAX - A_MIN)

        def obj(x):
            a_val = _inv_logit(x[0])
            sigma = 1e-8 + math.exp(float(x[1]))
            try:
                hw = ql.HullWhite(h, a_val, sigma)
                eng = ql.JamshidianSwaptionEngine(hw)
                err = 0.0
                for it in basket:
                    it["swpt"].setPricingEngine(eng)
                    model = float(it["swpt"].NPV())
                    mkt = float(it["mkt"])
                    err += (model - mkt) ** 2 / max(1.0, abs(mkt))
                return err
            except Exception:
                return 1e20

        # Initial guess: a=0.03, σ=0.005
        x0 = [_logit(0.03), math.log(0.005)]
        res = minimize(obj, x0, method="Nelder-Mead",
                       options={"maxiter": 1500, "xatol": 1e-8, "fatol": 1e-8})
        if not res.success:
            print(f"  [WARNING] Joint (a,σ) calibration did not converge: {res.message}")
        a_cal = _inv_logit(res.x[0])
        sigma_cal = 1e-8 + math.exp(float(res.x[1]))
        return a_cal, sigma_cal

    def _inverse_solve(self, target):
        def f(log_s):
            _, sw = self._build_berm()
            return self._price_berm(self.yts_h, sw, math.exp(log_s)) - target
        try:
            return math.exp(brentq(f, math.log(0.001), math.log(0.05), xtol=1e-12))
        except:
            res = minimize(lambda x: f(float(x[0]))**2, [math.log(0.007)],
                           method="Nelder-Mead", options={"maxiter": 300})
            return math.exp(float(res.x[0]))

    def calibrate(self):
        """Hybrid calibration: σ_ATM + Δσ_spread, optionally with joint (a, σ)."""
        self.basket = self._build_basket()

        if self.calib_a:
            print("\n  STEP 1: Joint (a, σ) calibration (European basket, ATM vols)")
            a_cal, sigma_cal = self._calib_joint(self.yts_h, self.basket)
            self.a = a_cal  # update a with calibrated value
            self.sigma_atm = sigma_cal
            print(f"    a_cal   = {a_cal:.6f}")
            print(f"    σ_ATM   = {sigma_cal:.6f} ({sigma_cal*10000:.2f} bp)")
        else:
            print("\n  STEP 1: σ_ATM calibration (a={:.4f} fixed, European basket, ATM vols)".format(self.a))
            self.sigma_atm = self._calib_sigma_atm(self.yts_h, self.basket)
            print(f"    σ_ATM   = {self.sigma_atm:.6f} ({self.sigma_atm*10000:.2f} bp)")

        _, berm_atm = self._build_berm()
        npv_atm = self._price_berm(self.yts_h, berm_atm, self.sigma_atm)

        if self.bbg_npv:
            diff_atm = 100.0 * (npv_atm - self.bbg_npv) / self.bbg_npv
            print(f"    NPV_ATM = {npv_atm:,.2f} ({diff_atm:+.1f}% vs BBG)")
        else:
            print(f"    NPV_ATM = {npv_atm:,.2f} (no BBG target)")

        if not self.bbg_npv:
            self.sigma_inv = self.sigma_atm
            self.delta_spread = 0.0
            self.sigma_total = self.sigma_atm
            _, self.berm = self._build_berm()
            self.npv = self._price_berm(self.yts_h, self.berm, self.sigma_total)
            return

        print(f"\n  STEP 2: σ_inverse (target NPV = {self.bbg_npv:,.2f})")
        self.sigma_inv = self._inverse_solve(self.bbg_npv)
        print(f"    σ_inv   = {self.sigma_inv:.6f} ({self.sigma_inv*10000:.2f} bp)")

        self.delta_spread = self.sigma_inv - self.sigma_atm
        self.sigma_total  = self.sigma_inv

        print(f"\n  STEP 3: Hybrid decomposition")
        if self.calib_a:
            print(f"    a       = {self.a:.6f} (calibrated)")
        else:
            print(f"    a       = {self.a:.6f} (fixed)")
        print(f"    σ_ATM     = {self.sigma_atm:.6f} ({self.sigma_atm*10000:.2f} bp)")
        print(f"    Δσ_spread = {self.delta_spread:.6f} ({self.delta_spread*10000:.2f} bp)")
        print(f"    σ_total   = {self.sigma_total:.6f} ({self.sigma_total*10000:.2f} bp)")

        # Final NPV
        _, self.berm = self._build_berm()
        self.npv = self._price_berm(self.yts_h, self.berm, self.sigma_total)

    def _bump_dfs(self, bp):
        return [d * math.exp(-bp/10000.0 * self.dc.yearFraction(self.val_date, dt))
                for dt, d in zip(self.node_dates, self.node_dfs)]

    def _reprice_with_dfs(self, ref, ev, dfs, sigma):
        saved = ql.Settings.instance().evaluationDate
        ql.Settings.instance().evaluationDate = ev
        try:
            h, _ = build_curve(ref, self.node_dates, dfs, self.cal, self.dc)
            ix = get_index(h, self.ccy)
            sc = make_schedule(self.swap_start, self.swap_end, self.fixed_tenor, self.cal, self.bdc)
            s = make_ois(self.direction, self.notional, sc, self.strike, ix,
                         self.fixed_dc, self.payment_lag, self.bdc, self.cal)
            s.setPricingEngine(ql.DiscountingSwapEngine(h))
            ed = [d for d in self.ex_dates if d > ev]
            if not ed:
                r = 0.0  # no exercise rights left → option expired → value = 0
            else:
                r = self._price_berm(h, make_swaption(s, ql.BermudanExercise(ed)), sigma)
            return r
        finally:
            ql.Settings.instance().evaluationDate = saved

    def _swap_npv_bumped(self, dfs):
        h, _ = build_curve(self.val_date, self.node_dates, dfs, self.cal, self.dc)
        ix = get_index(h, self.ccy)
        sc = make_schedule(self.swap_start, self.swap_end, self.fixed_tenor, self.cal, self.bdc)
        s = make_ois(self.direction, self.notional, sc, self.strike, ix,
                     self.fixed_dc, self.payment_lag, self.bdc, self.cal)
        s.setPricingEngine(ql.DiscountingSwapEngine(h))
        return float(s.NPV())

    def compute_greeks(self):
        """Compute all Greeks with hybrid Vega."""
        print("\n  Computing Greeks...")
        ref = self.val_date
        σ = self.sigma_total

        # DV01
        up = self._bump_dfs(+self.dv01_bp)
        dn = self._bump_dfs(-self.dv01_bp)
        pu = self._reprice_with_dfs(ref, ref, up, σ)
        pd = self._reprice_with_dfs(ref, ref, dn, σ)
        dv01 = (pd - pu) / (2.0 * self.dv01_bp)
        print(f"    DV01 done")

        # Gamma — dedicated bumps (gamma_bp may differ from dv01_bp)
        p0 = self.npv
        if abs(self.gamma_bp - self.dv01_bp) < 1e-12:
            # Same bump size: reuse DV01 repricing (no wasted compute)
            gamma = (pu - 2.0 * p0 + pd) / (self.dv01_bp ** 2)
        else:
            # Different bump: must reprice with gamma_bp
            gu = self._bump_dfs(+self.gamma_bp)
            gd = self._bump_dfs(-self.gamma_bp)
            pgu = self._reprice_with_dfs(ref, ref, gu, σ)
            pgd = self._reprice_with_dfs(ref, ref, gd, σ)
            gamma = (pgu - 2.0 * p0 + pgd) / (self.gamma_bp ** 2)
        print(f"    Gamma done")

        # Underlying DV01
        su = self._swap_npv_bumped(up)
        sd = self._swap_npv_bumped(dn)
        udv01 = (sd - su) / (2.0 * self.dv01_bp)

        # Delta — keep sign for hedge direction
        delta = (dv01 / udv01) if abs(udv01) > 1e-12 else 0.0
        print(f"    Delta done")

        # VEGA — hybrid
        def vega_bump(bp):
            bk = self._build_basket(vol_bump_bp=bp)
            if self.calib_a:
                _, sig_atm_b = self._calib_joint(self.yts_h, bk)
            else:
                sig_atm_b = self._calib_sigma_atm(self.yts_h, bk)
            sig_total_b = sig_atm_b + self.delta_spread
            _, sw = self._build_berm()
            return self._price_berm(self.yts_h, sw, sig_total_b)

        pv_up = vega_bump(+self.vega_bp)
        pv_dn = vega_bump(-self.vega_bp)
        vega = (pv_up - pv_dn) / (2.0 * self.vega_bp)
        print(f"    Vega done (hybrid)")

        # Theta — 1 calendar day roll (BBG convention)
        theta = 0.0
        if self.do_theta:
            nxt = ref + 1  # 1 calendar day, not business day
            dfn = self.yts_c.discount(nxt)
            tdfs = [self.yts_c.discount(d) / dfn for d in self.node_dates]
            pt = self._reprice_with_dfs(nxt, nxt, tdfs, σ)
            theta = pt - self.npv  # no annualization — BBG reports raw 1-day P&L
            print(f"    Theta done")

        self.greeks = dict(
            dv01=dv01, gamma_1bp=gamma, vega_1bp=vega,
            theta_1d=theta, delta_hedge=delta, underlying_dv01=udv01,
        )

    def print_results(self):
        """Print full results to console."""
        S = "=" * 90
        D = "-" * 90
        money = (self.strike - self.fair_rate) * 10000.0
        bps_leg = abs(float(self.swap.fixedLegBPS()))
        yv = self.npv / bps_leg if bps_leg else 0
        prem_pct = self.npv / self.notional * 100
        u_prem_pct = self.underlying_npv / self.notional * 100
        direction_str = "Receiver" if self.is_receiver else "Payer"

        print(f"\n{S}")
        print(f"BERMUDAN SWAPTION PRICER — {self.ccy} OIS + HW1F (v12 hybrid)")
        print(S)
        print(f"ValDate    : {self.val_date}")
        print(f"Deal       : {self.notional/1e6:.0f}MM {self.ccy} {direction_str}")
        print(f"Strike     : {self.strike*100:.6f}%")
        print(f"Swap       : {self.swap_start} → {self.swap_end}")
        print(f"ATM        : {self.fair_rate*100:.6f}%  |  Moneyness: {money:+.2f} bp OTM")
        print(f"Exercises  : {len(self.ex_dates)} dates")
        print(f"Model      : HW1F | a={self.a} | FDM {self.fdm_t}×{self.fdm_x}")
        print(f"σ_ATM={self.sigma_atm:.6f} + Δσ={self.delta_spread:.6f} → σ_total={self.sigma_total:.6f}")

        print(f"\n{S}")
        print("RESULTS")
        print(S)
        print(f"NPV              : {self.npv:>14,.2f} {self.ccy}")
        print(f"Yield Value      : {yv:>14.3f} bps")
        print(f"Premium          : {prem_pct:>14.5f}%")
        print(f"Underlying Prem  : {u_prem_pct:>14.5f}%")
        print(f"Underlying NPV   : {self.underlying_npv:>14,.2f} {self.ccy}")

        print(f"\n{S}")
        print("GREEKS")
        print(S)
        g = self.greeks
        print(f"DV01             : {g['dv01']:>14,.2f}")
        print(f"Gamma (1bp)      : {g['gamma_1bp']:>14,.2f}")
        print(f"Vega (1bp)       : {g['vega_1bp']:>14,.2f}")
        print(f"Theta (1-day)    : {g['theta_1d']:>14,.2f}")
        print(f"Delta (Hedge)    : {g['delta_hedge']:>14.5f}")
        print(f"Underlying DV01  : {g['underlying_dv01']:>14,.2f}")

        # BBG comparison if available
        b = self.bbg
        if b.get("dv01") or b.get("vega_1bp"):
            print(f"\n{S}")
            print("BBG COMPARISON")
            print(S)
            print(f"{'Metric':<22} {'Bloomberg':>15} {'QuantLib':>15} {'Diff':>15}")
            print(D)

            nd = 100*(self.npv-self.bbg_npv)/self.bbg_npv if self.bbg_npv else 0
            print(f"{'NPV':22} {self.bbg_npv:>15,.2f} {self.npv:>15,.2f} {nd:>14.2f}%")

            if b.get("atm_strike"):
                ad = (self.fair_rate - float(b["atm_strike"])/100) * 10000
                print(f"{'ATM (%)':22} {float(b['atm_strike']):>15.6f} {self.fair_rate*100:>15.6f} {ad:>13.2f} bp")
            if b.get("yield_value_bp"):
                print(f"{'Yield Value (bp)':22} {float(b['yield_value_bp']):>15.3f} {yv:>15.3f} {yv-float(b['yield_value_bp']):>13.3f} bp")

            for label, k_g, k_b in [
                ("DV01",           "dv01",           "dv01"),
                ("Gamma (1bp)",    "gamma_1bp",      "gamma_1bp"),
                ("Vega (1bp)",     "vega_1bp",       "vega_1bp"),
                ("Theta (1-day)",  "theta_1d",       "theta_1d"),
                ("Delta (Hedge)",  "delta_hedge",    "delta_hedge"),
                ("Underlying DV01","underlying_dv01","underlying_dv01"),
            ]:
                bv = b.get(k_b)
                if bv is not None:
                    bv = float(bv)
                    qv = float(g[k_g])
                    if k_g == "delta_hedge":
                        print(f"{label:22} {bv:>15.5f} {qv:>15.5f} {qv-bv:>15.5f}")
                    else:
                        print(f"{label:22} {bv:>15,.2f} {qv:>15,.2f} {qv-bv:>15,.2f}")
            print(S)

    def export_excel(self, filepath):
        """Export results to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("  [WARNING] openpyxl not installed — skipping Excel export")
            print("  Install with: pip install openpyxl")
            return

        wb = openpyxl.Workbook()

        # --- Sheet 1: Results ---
        ws = wb.active
        ws.title = "Results"

        header_font = Font(bold=True, size=12)
        label_font  = Font(bold=True)
        num_fmt     = '#,##0.00'
        pct_fmt     = '0.00000%'

        row = 1
        ws.cell(row, 1, "Bermudan Swaption Pricer — Results").font = Font(bold=True, size=14)
        row += 2

        # Deal info
        ws.cell(row, 1, "DEAL").font = header_font; row += 1
        deal_info = [
            ("Valuation Date", str(self.val_date)),
            ("Notional", self.notional),
            ("Strike", f"{self.strike*100:.6f}%"),
            ("Direction", "Receiver" if self.is_receiver else "Payer"),
            ("Swap", f"{self.swap_start} → {self.swap_end}"),
            ("ATM Rate", f"{self.fair_rate*100:.6f}%"),
            ("Moneyness", f"{(self.strike-self.fair_rate)*10000:+.2f} bp"),
            ("Exercise Dates", len(self.ex_dates)),
        ]
        for label, val in deal_info:
            ws.cell(row, 1, label).font = label_font
            ws.cell(row, 2, val)
            row += 1

        row += 1
        ws.cell(row, 1, "MODEL").font = header_font; row += 1
        model_info = [
            ("σ_ATM", f"{self.sigma_atm:.6f} ({self.sigma_atm*10000:.2f} bp)"),
            ("Δσ_spread", f"{self.delta_spread:.6f} ({self.delta_spread*10000:.2f} bp)"),
            ("σ_total", f"{self.sigma_total:.6f} ({self.sigma_total*10000:.2f} bp)"),
            ("Mean Reversion (a)", self.a),
            ("FDM Grid", f"{self.fdm_t}×{self.fdm_x}"),
        ]
        for label, val in model_info:
            ws.cell(row, 1, label).font = label_font
            ws.cell(row, 2, val)
            row += 1

        row += 1
        ws.cell(row, 1, "VALUATION").font = header_font; row += 1
        bps_leg = abs(float(self.swap.fixedLegBPS()))
        yv = self.npv / bps_leg if bps_leg else 0
        val_info = [
            ("NPV", f"{self.npv:,.2f}"),
            ("Yield Value (bp)", f"{yv:.3f}"),
            ("Premium (%)", f"{self.npv/self.notional*100:.5f}"),
            ("Underlying NPV", f"{self.underlying_npv:,.2f}"),
        ]
        for label, val in val_info:
            ws.cell(row, 1, label).font = label_font
            ws.cell(row, 2, val)
            row += 1

        row += 1
        ws.cell(row, 1, "GREEKS").font = header_font; row += 1
        g = self.greeks
        greek_info = [
            ("DV01", f"{g['dv01']:,.2f}"),
            ("Gamma (1bp)", f"{g['gamma_1bp']:,.2f}"),
            ("Vega (1bp)", f"{g['vega_1bp']:,.2f}"),
            ("Theta (1-day)", f"{g['theta_1d']:,.2f}"),
            ("Delta (Hedge)", f"{g['delta_hedge']:.5f}"),
            ("Underlying DV01", f"{g['underlying_dv01']:,.2f}"),
        ]
        for label, val in greek_info:
            ws.cell(row, 1, label).font = label_font
            ws.cell(row, 2, val)
            row += 1

        # BBG comparison
        b = self.bbg
        if b.get("dv01") or b.get("vega_1bp"):
            row += 1
            ws.cell(row, 1, "BBG COMPARISON").font = header_font; row += 1
            ws.cell(row, 1, "Metric").font = label_font
            ws.cell(row, 2, "Bloomberg").font = label_font
            ws.cell(row, 3, "QuantLib").font = label_font
            ws.cell(row, 4, "Diff").font = label_font
            row += 1

            comps = [
                ("NPV", self.bbg_npv, self.npv),
                ("DV01", b.get("dv01"), g["dv01"]),
                ("Gamma", b.get("gamma_1bp"), g["gamma_1bp"]),
                ("Vega", b.get("vega_1bp"), g["vega_1bp"]),
                ("Theta", b.get("theta_1d"), g["theta_1d"]),
                ("Delta", b.get("delta_hedge"), g["delta_hedge"]),
                ("Und. DV01", b.get("underlying_dv01"), g["underlying_dv01"]),
            ]
            for label, bv, qv in comps:
                ws.cell(row, 1, label)
                if bv is not None:
                    ws.cell(row, 2, float(bv))
                    ws.cell(row, 3, float(qv))
                    ws.cell(row, 4, float(qv) - float(bv))
                else:
                    ws.cell(row, 2, "N/A")
                    ws.cell(row, 3, float(qv))
                row += 1

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

        # --- Sheet 2: Curve ---
        ws2 = wb.create_sheet("Curve")
        ws2.cell(1, 1, "Date").font = label_font
        ws2.cell(1, 2, "Discount Factor").font = label_font
        for i, (d, df) in enumerate(self.mkt["curve"], 2):
            ws2.cell(i, 1, str(d))
            ws2.cell(i, 2, float(df))

        # --- Sheet 3: Vol Surface ---
        ws3 = wb.create_sheet("Vol Surface")
        ws3.cell(1, 1, "Expiry \\ Tenor").font = label_font
        vsd = self.cfg.get("vol_surface_data", {})
        tnr_labels = vsd.get("tenor_labels", [f"{t:.0f}Y" for t in self.tnr_grid])
        exp_labels = vsd.get("expiry_labels", [f"{e:.2f}" for e in self.exp_grid])
        for j, t in enumerate(tnr_labels):
            ws3.cell(1, j+2, t).font = label_font
        raw_vol = self.vol_mat * 1000.0  # back to BPx10
        for i, exp in enumerate(exp_labels):
            ws3.cell(i+2, 1, exp)
            for j in range(raw_vol.shape[1]):
                ws3.cell(i+2, j+2, float(raw_vol[i, j]))

        wb.save(filepath)
        print(f"  Results exported to: {filepath}")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Bermudan Swaption Pricer (v12)")
    parser.add_argument("--config", default=None, help="Path to YAML config file")
    parser.add_argument("--output", default=None, help="Override Excel output path")
    args = parser.parse_args()

    # Load config
    config_path = args.config
    if config_path is None:
        # Search in common locations
        candidates = [
            "config.yaml",
            os.path.join("config", "config.yaml"),
            os.path.join(os.path.dirname(__file__), "..", "config", "config.yaml"),
        ]
        for c in candidates:
            if os.path.exists(c):
                config_path = c
                break
        if config_path is None:
            print("Config file not found. Searched:")
            for c in candidates:
                print(f"  - {c}")
            print("\nUse: python pricer.py --config path/to/config.yaml")
            sys.exit(1)

    config_dir = os.path.dirname(os.path.abspath(config_path))

    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    print("=" * 90)
    print("BERMUDAN SWAPTION PRICER (v12 hybrid)")
    print("=" * 90)
    print(f"Config: {config_path}")

    # Fetch data
    print("\n[1/4] FETCHING MARKET DATA")
    mkt = fetch_all(cfg, config_dir=config_dir)

    # Build pricer
    print("\n[2/4] CALIBRATION")
    pricer = BermudanPricer(cfg, mkt)
    pricer.setup()
    pricer.calibrate()

    # Greeks
    print("\n[3/4] GREEKS")
    pricer.compute_greeks()

    # Output
    print("\n[4/4] OUTPUT")
    pricer.print_results()

    out_cfg = cfg.get("output", {})
    if out_cfg.get("export_excel", False):
        xlsx_path = args.output or out_cfg.get("excel_file", "bermudan_results.xlsx")
        if not os.path.isabs(xlsx_path):
            xlsx_path = os.path.join(config_dir, xlsx_path)
        pricer.export_excel(xlsx_path)

    print("\n✓ Done")


if __name__ == "__main__":
    main()

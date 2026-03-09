#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
excel_bridge.py — Excel ↔ Pricer Bridge
=========================================
Reads deal parameters from an Excel workbook ("Deal" sheet),
runs the Bermudan pricer, writes results back to a "Results" sheet.

Usage:
  python excel_bridge.py deal_book.xlsx
  python excel_bridge.py deal_book.xlsx --sheet "MyDeal"

Expected Excel layout (sheet "Deal" or custom name):
  Column A: parameter names
  Column B: values

Required rows (case-insensitive match):
  valuation_date     | 2026-01-30
  notional           | 10000000
  strike             | 3.14817          (in %)
  direction          | Receiver
  swap_start         | 2027-02-12
  swap_end           | 2032-02-12
  frequency          | SemiAnnual
  day_count          | ACT/365
  payment_lag        | 2
  currency           | CAD
  mean_reversion     | 0.03
  fdm_grid           | 300
  bbg_npv            | 202935.23
  data_mode          | manual           (or bloomberg)

Optional rows:
  exercise_mode      | auto             (or custom)
  exercise_dates     | 2027-02-12, 2027-08-12, ...  (comma-separated)
  bbg_atm            | 3.006868
  bbg_dv01           | 2464.17
  bbg_gamma          | 21.58
  bbg_vega           | 2670.37
  bbg_theta          | -125.53
  bbg_delta          | 0.61381
  bbg_udv01          | 4614.51
  bbg_yield_value    | 44.879
  curve_sheet        | Curve            (name of sheet with curve data)
  vol_sheet          | VolSurface       (name of sheet with vol data)

Curve sheet layout:
  Column A: dates (YYYY-MM-DD)
  Column B: discount factors

Vol surface sheet layout:
  Row 1: header — blank, then tenor labels (1Y, 2Y, ...)
  Col A: expiry labels (1Mo, 3Mo, ...)
  Values: BPx10
"""

import argparse
import os
import sys
import yaml
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)


def read_deal_sheet(wb, sheet_name="Deal"):
    """Read parameter=value pairs from Excel sheet."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    params = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] is not None:
            key = str(row[0]).strip().lower().replace(" ", "_")
            val = row[1]
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            params[key] = val
    return params


def read_curve_sheet(wb, sheet_name="Curve"):
    """Read curve data from Excel: col A=date, col B=DF."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] is None:
            continue
        d = row[0]
        if isinstance(d, datetime):
            d = d.strftime("%Y-%m-%d")
        d_str = str(d).strip()
        # Skip header
        if d_str.lower() in ("date", "dates", "tenor", ""):
            continue
        try:
            df = float(row[1])
            data.append([d_str, df])
        except (TypeError, ValueError):
            continue
    return data


def read_vol_sheet(wb, sheet_name="VolSurface"):
    """Read vol surface from Excel. Returns (values, expiry_labels, tenor_labels)."""
    if sheet_name not in wb.sheetnames:
        return None, None, None

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None, None

    # First row: header with tenor labels (skip first cell)
    tenor_labels = [str(c).strip() for c in rows[0][1:] if c is not None]

    expiry_labels = []
    values = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        expiry_labels.append(str(row[0]).strip())
        vals = []
        for c in row[1:1+len(tenor_labels)]:
            try:
                vals.append(float(c))
            except (TypeError, ValueError):
                vals.append(0.0)
        values.append(vals)

    return values, expiry_labels, tenor_labels


def build_config(params, curve_data, vol_data):
    """Build config dict from parsed Excel data."""
    p = params

    def get(key, default=None):
        return p.get(key, default)

    cfg = {
        "deal": {
            "valuation_date": str(get("valuation_date", "2026-01-30")),
            "notional": float(get("notional", 10_000_000)),
            "strike": float(get("strike", 3.0)),
            "direction": str(get("direction", "Receiver")),
            "swap_start": str(get("swap_start", "2027-02-12")),
            "swap_end": str(get("swap_end", "2032-02-12")),
            "fixed_frequency": str(get("frequency", "SemiAnnual")),
            "day_count": str(get("day_count", "ACT/365")),
            "payment_lag": int(float(get("payment_lag", 2))),
            "currency": str(get("currency", "CAD")),
        },
        "exercise": {
            "mode": str(get("exercise_mode", "auto")),
        },
        "model": {
            "name": "HW1F",
            "mean_reversion": float(get("mean_reversion", 0.03)),
            "calibrate_a": False,
            "fdm_time_grid": int(float(get("fdm_grid", 300))),
            "fdm_space_grid": int(float(get("fdm_grid", 300))),
        },
        "greeks": {
            "dv01_bump_bp": 1.0,
            "gamma_bump_bp": 1.0,
            "vega_bump_bp": 1.0,
            "compute_theta": True,
            "theta_annualization": "365/252",
        },
        "data_source": {
            "mode": str(get("data_mode", "manual")),
            "manual": {},
            "bloomberg": {
                "curve_ticker": str(get("bbg_curve_ticker", "YCSW0147 Index")),
                "timeout_ms": 30000,
            },
        },
        "benchmark": {
            "npv": float(get("bbg_npv", 0)),
        },
        "output": {
            "print_console": True,
            "export_excel": True,
            "excel_file": "bermudan_results.xlsx",
        },
    }

    # Exercise dates
    if get("exercise_dates"):
        dates_str = str(get("exercise_dates"))
        cfg["exercise"]["custom_dates"] = [d.strip() for d in dates_str.split(",")]
        cfg["exercise"]["mode"] = "custom"

    # BBG benchmark Greeks (optional)
    for key, cfg_key in [
        ("bbg_atm", "atm_strike"),
        ("bbg_dv01", "dv01"),
        ("bbg_gamma", "gamma_1bp"),
        ("bbg_vega", "vega_1bp"),
        ("bbg_theta", "theta_1d"),
        ("bbg_delta", "delta_hedge"),
        ("bbg_udv01", "underlying_dv01"),
        ("bbg_yield_value", "yield_value_bp"),
    ]:
        v = get(key)
        if v is not None:
            try:
                cfg["benchmark"][cfg_key] = float(v)
            except (TypeError, ValueError):
                pass

    # Curve data
    if curve_data:
        cfg["curve_data"] = curve_data

    # Vol surface data
    if vol_data:
        values, exp_labels, tnr_labels = vol_data
        if values:
            cfg["vol_surface_data"] = {
                "expiry_labels": exp_labels,
                "tenor_labels": tnr_labels,
                "values": values,
            }

    return cfg


def main():
    parser = argparse.ArgumentParser(description="Excel → Pricer Bridge")
    parser.add_argument("workbook", help="Path to Excel workbook with deal parameters")
    parser.add_argument("--sheet", default="Deal", help="Sheet name with deal params (default: Deal)")
    parser.add_argument("--curve-sheet", default="Curve", help="Sheet with curve data")
    parser.add_argument("--vol-sheet", default="VolSurface", help="Sheet with vol surface")
    parser.add_argument("--output", default=None, help="Output Excel file path")
    args = parser.parse_args()

    if not os.path.exists(args.workbook):
        print(f"File not found: {args.workbook}")
        sys.exit(1)

    print("=" * 90)
    print("EXCEL → BERMUDAN PRICER BRIDGE")
    print("=" * 90)

    # Read Excel
    print(f"\nReading: {args.workbook}")
    wb = openpyxl.load_workbook(args.workbook, data_only=True)

    params = read_deal_sheet(wb, args.sheet)
    print(f"  Deal parameters: {len(params)} entries")

    curve_data = read_curve_sheet(wb, args.curve_sheet)
    if curve_data:
        print(f"  Curve: {len(curve_data)} nodes from '{args.curve_sheet}' sheet")

    vol_values, vol_exp, vol_tnr = read_vol_sheet(wb, args.vol_sheet)
    vol_data = (vol_values, vol_exp, vol_tnr) if vol_values else None
    if vol_data:
        print(f"  Vol surface: {len(vol_values)}×{len(vol_tnr)} from '{args.vol_sheet}' sheet")

    wb.close()

    # Build config
    cfg = build_config(params, curve_data, vol_data)

    # Write temp config
    config_dir = os.path.dirname(os.path.abspath(args.workbook))
    config_path = os.path.join(config_dir, "_temp_config.yaml")
    with open(config_path, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, default_flow_style=False, allow_unicode=True)
    print(f"  Config written to: {config_path}")

    # Run pricer
    from pricer import BermudanPricer
    from bbg_fetcher import fetch_all

    print("\n" + "=" * 90)
    print("RUNNING PRICER")
    print("=" * 90)

    mkt = fetch_all(cfg, config_dir=config_dir)
    pricer = BermudanPricer(cfg, mkt)
    pricer.setup()
    pricer.calibrate()
    pricer.compute_greeks()
    pricer.print_results()

    # Export
    xlsx_out = args.output or os.path.join(config_dir, "bermudan_results.xlsx")
    pricer.export_excel(xlsx_out)

    # Cleanup temp config
    try:
        os.remove(config_path)
    except:
        pass

    print("\n✓ Done")


if __name__ == "__main__":
    main()
